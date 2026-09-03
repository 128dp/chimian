"""
Excel file generation for material metrics
Extensible design allowing custom columns, metrics, and formatting
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from io import BytesIO
from typing import Dict, List, Optional, Callable, Any
from collections import defaultdict

class ExcelWriter:
    """
    Write material metrics to Excel file with standardized units.
    Extensible design supports custom columns, metrics, and formatting.
    """
    
    # Default column configuration
    DEFAULT_COLUMNS = [
        {'name': 'Company', 'width': 18, 'format': None},
        {'name': 'Tensile Strength (MPa)', 'width': 18, 'format': '0.0000'},
        {'name': 'Flexural Strength (MPa)', 'width': 20, 'format': '0.0000'},
        {'name': 'Density (tonnes/mm³)', 'width': 18, 'format': '0.0000e+00'},
        {'name': 'Tensile Modulus (MPa)', 'width': 19, 'format': '0.0000'},
        {'name': 'Flexural Modulus (MPa)', 'width': 20, 'format': '0.0000'},
        {'name': 'Elongation (strain)', 'width': 18, 'format': '0.0000'},
    ]
    
    def __init__(self, output_path: str, columns: Optional[List[Dict]] = None, 
                 include_test_method: bool = True,
                 existing_workbook_bytes: Optional[bytes] = None,
                 existing_workbook_path: Optional[str] = None):
        """
        Args:
            output_path: Where the final workbook will be saved.
            existing_workbook_bytes: Optional bytes of an existing .xlsx file
                to merge new results into. Existing sheets/materials are kept
                as-is; new material tabs become new sheets, and materials
                that already have a sheet get their new companies appended
                below the existing content.
            existing_workbook_path: Same as above, but as a file path instead
                of raw bytes.
        """
        self.output_path = output_path

        if existing_workbook_bytes is not None:
            self.wb = openpyxl.load_workbook(BytesIO(existing_workbook_bytes))
        elif existing_workbook_path:
            self.wb = openpyxl.load_workbook(existing_workbook_path)
        else:
            self.wb = openpyxl.Workbook()
            self.wb.remove(self.wb.active)  # Remove default sheet

        self.materials_data = defaultdict(list)  # {material_name: [records]}
        
        # Use custom columns or defaults
        self.columns = columns if columns else self.DEFAULT_COLUMNS
        
        # Filter out test method column if not needed
        if not include_test_method:
            self.columns = [c for c in self.columns if c['name'] != 'Test Method']
        
        self.include_test_method = include_test_method
        self.custom_formatters = {}  # {column_name: formatter_function}
    
    def add_custom_formatter(self, column_name: str, formatter: Callable[[Any], Any]) -> None:
        """
        Add a custom formatter function for a column.
        
        Args:
            column_name: Name of the column to format
            formatter: Function that takes a value and returns formatted value
        """
        self.custom_formatters[column_name] = formatter
    
    def add_custom_column(self, column_config: Dict) -> None:
        """
        Add a custom column to the default columns.
        
        Args:
            column_config: Dict with 'name', 'width', and optional 'format'
        """
        if 'name' in column_config:
            self.columns.append(column_config)
    
    def remove_column(self, column_name: str) -> None:
        """Remove a column by name"""
        self.columns = [c for c in self.columns if c['name'] != column_name]
    
    def set_custom_columns(self, columns: List[Dict]) -> None:
        """Replace all columns with custom configuration"""
        self.columns = columns
    
    def add_record(self, **kwargs):
        """
        Add a single metric record. Supports both old and new formats.
        
        New format (all 6 metrics in one record):
            add_record(company='Kingfa', material='ABS', 
                      tensile_strength_original_value=50.0,
                      tensile_strength_standard_value=50.0, ...)
        
        Old format (one metric per record):
            add_record(company='Kingfa', material='ABS', metric_name='Tensile Strength',
                      original_value=50.0, original_unit='MPa', 
                      standard_value=50.0, standard_unit='MPa')
        
        Args:
            company: Manufacturer name
            material: Material type
            **kwargs: Metric fields (flexible format)
        """
        if not kwargs.get('company') or not kwargs.get('material'):
            return  # Skip invalid records
        
        # Create record from all kwargs
        record = dict(kwargs)
        
        # Store the record under the material key
        material = record.get('material')
        if material:
            self.materials_data[material].append(record)
    
    def create_sheets(self):
        """Create Excel sheets organized by material"""
        for material, records in self.materials_data.items():
            self._create_material_sheet(material, records)
    
    def _create_material_sheet(self, material: str, records: List[Dict]):
        """Create a single sheet for a material with company sections (vertical layout)"""
        # Sanitize sheet name (Excel has 31 character limit)
        sheet_name = material[:31] if material else "Material"
        
        # If this material already has a sheet (e.g. merging into an
        # existing workbook), append new company sections below the
        # existing content instead of overwriting the sheet.
        if sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            current_row = ws.max_row + 2  # leave a blank separator row
        else:
            ws = self.wb.create_sheet(title=sheet_name)
            current_row = 1
        
        # Define styles
        company_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        company_font = Font(bold=True, color="FFFFFF", size=12)
        
        metric_header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        metric_header_font = Font(bold=True, size=11, color="000000")
        
        metric_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        value_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        data_font = Font(color="000000")
        
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        num_columns = 7
        last_col_letter = get_column_letter(num_columns)
        
        # Metric names mapping
        metric_names = {
            'tensile_strength': 'Tensile Strength',
            'flexural_strength': 'Flexural Strength',
            'density': 'Density',
            'tensile_modulus': 'Tensile Modulus',
            'flexural_modulus': 'Flexural Modulus',
            'elongation': 'Elongation'
        }
        
        metric_order = [
            'tensile_strength',
            'flexural_strength',
            'density',
            'tensile_modulus',
            'flexural_modulus',
            'elongation'
        ]
        
        # Add each company as a section
        for record in records:
            company_name = record.get('company', 'Unknown')
            metrics = record.get('metrics', {})
            
            # Company header row
            company_cell = ws.cell(row=current_row, column=1)
            company_cell.value = company_name
            company_cell.fill = company_fill
            company_cell.font = company_font
            company_cell.alignment = Alignment(horizontal='left', vertical='center')
            company_cell.border = border
            ws.merge_cells(f'A{current_row}:{last_col_letter}{current_row}')
            
            current_row += 1
            
            # Metric header row
            headers = ['PROPERTY', 'TEST METHOD', 'TEST CONDITION', 'UNIT (FILE)', 'VALUE (FILE)', 'UNIT (SI)', 'VALUE (SI)']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = header
                cell.fill = metric_header_fill
                cell.font = metric_header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            current_row += 1
            
            # Metric rows - a property may have multiple entries when the
            # datasheet lists it under different test conditions
            for metric_key in metric_order:
                entries = metrics.get(metric_key) or [{}]
                property_name = metric_names.get(metric_key, metric_key)
                
                for entry in entries:
                    if not isinstance(entry, dict):
                        # Defensive: guard against malformed (non-dict) entries
                        entry = {}
                    
                    row_values = [
                        property_name,
                        entry.get('test_method') or '-',
                        entry.get('test_condition') or '-',
                        entry.get('unit_file') or '-',
                        entry.get('value_file') if entry.get('value_file') not in (None, '') else '-',
                        entry.get('unit_si') or '-',
                        entry.get('value_si') if entry.get('value_si') not in (None, '') else '-',
                    ]
                    
                    for col_num, val in enumerate(row_values, 1):
                        cell = ws.cell(row=current_row, column=col_num)
                        cell.value = val
                        cell.fill = metric_fill if col_num % 2 == 1 else value_fill
                        cell.font = data_font
                        cell.alignment = Alignment(
                            horizontal='left' if col_num == 1 else 'center',
                            vertical='center'
                        )
                        cell.border = border
                    
                    current_row += 1
            
            # Blank row between companies
            current_row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 22  # Property
        ws.column_dimensions['B'].width = 16  # Test Method
        ws.column_dimensions['C'].width = 30  # Test Condition
        ws.column_dimensions['D'].width = 14  # Unit (file)
        ws.column_dimensions['E'].width = 16  # Value (file)
        ws.column_dimensions['F'].width = 16  # Unit (SI)
        ws.column_dimensions['G'].width = 18  # Value (SI)
    
    def save(self):
        """Save the workbook"""
        self.wb.save(self.output_path)
        print(f"Excel file saved to: {self.output_path}")


class DataOrganizer:
    """Organize extracted data and prepare for Excel export"""
    
    METRIC_NAMES = {
        'tensile_strength': 'Tensile Strength',
        'flexural_strength': 'Flexural Strength',
        'density': 'Density',
        'tensile_modulus': 'Tensile Modulus',
        'flexural_modulus': 'Flexural Modulus',
        'elongation': 'Elongation'
    }
    
    def __init__(self):
        self.excel_writer = None
        self.extracted_data = []
    
    def add_extraction_result(self, result: Dict):
        """Add an extraction result from PDF"""
        self.extracted_data.append(result)
    
    def generate_excel(self, output_path: str, unit_converter):
        """Generate Excel file from all extracted data"""
        self.excel_writer = ExcelWriter(output_path)
        
        for extraction in self.extracted_data:
            company = extraction.get('company', 'Unknown')
            material = extraction.get('material', 'Unknown')
            metrics = extraction.get('metrics', {})
            
            for metric_key, (original_value, original_unit) in metrics.items():
                metric_name = self.METRIC_NAMES.get(metric_key, metric_key)
                
                # Convert to standard units
                if metric_key == 'tensile_strength' or metric_key == 'flexural_strength':
                    std_value, std_unit = unit_converter.convert_stress(original_value, original_unit)
                
                elif metric_key == 'density':
                    std_value, std_unit = unit_converter.convert_density(original_value, original_unit)
                
                elif metric_key == 'tensile_modulus' or metric_key == 'flexural_modulus':
                    std_value, std_unit = unit_converter.convert_stress(original_value, original_unit)
                
                elif metric_key == 'elongation':
                    std_value, std_unit = unit_converter.convert_elongation(original_value, original_unit)
                
                else:
                    std_value, std_unit = original_value, original_unit
                
                self.excel_writer.add_record(
                    company=company,
                    material=material,
                    metric_name=metric_name,
                    original_value=original_value,
                    original_unit=original_unit,
                    standard_value=std_value,
                    standard_unit=std_unit,
                    test_method=""
                )
        
        self.excel_writer.create_sheets()
        self.excel_writer.save()
